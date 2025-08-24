import os, sys, builtins
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier

def letra_cabecalho(cabecalho: str, ws: str):
    coluna = None
    for col in range(1, ws.max_column + 1):
        cell_value = (ws[f"{get_column_letter(col)}1"].value).upper()
        if cell_value == cabecalho.upper():
            coluna = col
            print(f"Cabeçalho {cabecalho} da coluna {coluna}")
            return get_column_letter(coluna)
    
    print(f"Não existe o cabeçalho {cabecalho}")
    return coluna

def resource_path(relative_path):
    """Obtem o caminho absoluto para o recurso, funciona para dev e PyInstaller."""
    try:
        # PyInstaller cria uma pasta temporária e define essa variável
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#Excel que vai ser modificado
path_excel = os.path.join(r"C:\Users\Prime Contabil\Downloads", "TW CHECK 2025 07.xlsx")
wb = load_workbook(path_excel)
ws_compras_produtos = wb["COMPRAS PRODUTOS"]
ws_vendas_produtos = wb["VENDAS PRODUTOS"]

#Criando uma nova tabela
if letra_cabecalho("TIPO CFOP", ws_compras_produtos) == None:
    ws_compras_produtos[get_column_letter(ws_compras_produtos.max_column + 1) + "1"] = "TIPO CFOP"
    ws_compras_produtos[get_column_letter(ws_compras_produtos.max_column + 1) + "1"] = "TIPO CLASSIFICAÇÃO"

#Criando uma nova tabela
if letra_cabecalho("TIPO CFOP", ws_vendas_produtos) == None:
    ws_vendas_produtos[get_column_letter(ws_vendas_produtos.max_column + 1) + "1"] = "TIPO CFOP"
    ws_vendas_produtos[get_column_letter(ws_vendas_produtos.max_column + 1) + "1"] = "TIPO CLASSIFICAÇÃO"

#Pegando a letra da coluna
coluna_compras_cfop = letra_cabecalho("CFOP", ws_compras_produtos)
coluna_compras_ncm = letra_cabecalho("Classificação", ws_compras_produtos)
coluna_compras_cest = letra_cabecalho("CEST", ws_compras_produtos)
coluna_compras_produto = letra_cabecalho("NOME PRODUTO", ws_compras_produtos)
coluna_compras_tipo_cfop = letra_cabecalho("TIPO CFOP", ws_compras_produtos)
coluna_compras_pis = letra_cabecalho("CST PIS", ws_compras_produtos)
coluna_compras_cofins = letra_cabecalho("CST COFINS", ws_compras_produtos)
coluna_compras_icms = letra_cabecalho("CST ICMS", ws_compras_produtos)

#Criando a aba PRODUTOS e copiado os produtos para la
wb.create_sheet("PRODUTOS")
aba_produtos = wb["PRODUTOS"]
aba_produtos["A1"] = "PRODUTOS"
aba_produtos["B1"] = "NCM"
aba_produtos["C1"] = "CFOP"
aba_produtos["D1"] = "ID-CFOP"
aba_produtos["E1"] = "CLASSIFICAÇÃO"
aba_produtos["F1"] = "CONFIANCA"
aba_produtos["G1"] = "CEST PIS"
aba_produtos["H1"] = "CEST COFINS"
aba_produtos["I1"] = "CEST ICMS"
#aba_produtos["J1"] = "CORREÇÃO PIS"
#aba_produtos["K1"] = "CORREÇÃO COFINS"
#aba_produtos["L1"] = "CORREÇÃO ICMS"

#Pegando os produtos, ncm e cfop
produtos_unicos = set()
row_destino = 2

#Iterar linha por linha para manter a relação produto-NCM-CFOP NO COMPRAS
for row_num in range(2, ws_compras_produtos.max_row + 1):
    produto = ws_compras_produtos[f"{coluna_compras_produto}{row_num}"].value
    ncm = ws_compras_produtos[f"{coluna_compras_ncm}{row_num}"].value
    cfop = ws_compras_produtos[f"{coluna_compras_cfop}{row_num}"].value
    pis = ws_compras_produtos[f"{coluna_compras_pis}{row_num}"].value
    cofins = ws_compras_produtos[f"{coluna_compras_cofins}{row_num}"].value
    icms = ws_compras_produtos[f"{coluna_compras_icms}{row_num}"].value
    
    #Criar uma chave única para evitar duplicatas
    chave_produto = (produto, ncm, cfop)
    
    if produto is not None and chave_produto not in produtos_unicos:
        produtos_unicos.add(chave_produto)
        
        #Escrever na planilha destino
        aba_produtos[f"A{row_destino}"] = produto
        aba_produtos[f"B{row_destino}"] = ncm
        aba_produtos[f"C{row_destino}"] = cfop
        aba_produtos[f"G{row_destino}"] = pis
        aba_produtos[f"H{row_destino}"] = cofins
        aba_produtos[f"I{row_destino}"] = icms
        row_destino += 1

#Iterar linha por linha para manter a relação produto-NCM-CFOP NO VENDAS
for row_num in range(2, ws_vendas_produtos.max_row + 1):
    produto = ws_vendas_produtos[f"{coluna_compras_produto}{row_num}"].value
    ncm = ws_vendas_produtos[f"{coluna_compras_ncm}{row_num}"].value
    cfop = ws_vendas_produtos[f"{coluna_compras_cfop}{row_num}"].value
    pis = ws_vendas_produtos[f"{coluna_compras_pis}{row_num}"].value
    cofins = ws_vendas_produtos[f"{coluna_compras_cofins}{row_num}"].value
    icms = ws_vendas_produtos[f"{coluna_compras_icms}{row_num}"].value
    
    #Criar uma chave única para evitar duplicatas
    chave_produto = (produto, ncm, cfop)
    
    if produto is not None and chave_produto not in produtos_unicos:
        produtos_unicos.add(chave_produto)
        
        #Escrever na planilha destino
        aba_produtos[f"A{row_destino}"] = produto
        aba_produtos[f"B{row_destino}"] = ncm
        aba_produtos[f"C{row_destino}"] = cfop
        aba_produtos[f"G{row_destino}"] = pis
        aba_produtos[f"H{row_destino}"] = cofins
        aba_produtos[f"I{row_destino}"] = icms
        row_destino += 1

#ws_produtos.insert_rows(idx=2, amount=10)
wb.save("TESTE.xlsx")
#Ate aqui OK

#Carregar a base tributaria
df_matriz_tributaria = pd.read_excel('MatrizTributaria.xlsx')
X_train = df_matriz_tributaria['PRODUTO'].astype(str)
y_train = df_matriz_tributaria['CLASSIFICAÇÃO'].astype(str)

df_cfop = pd.read_excel('MatrizTributaria.xlsx', sheet_name="CFOP")
cfop_series = df_cfop['CFOP'].astype(int)

cfop_numeric = pd.to_numeric(cfop_series, errors='coerce')
df_cfop['_CFOP_num'] = cfop_numeric

cfop_to_id = dict(zip(df_cfop['CFOP'], df_cfop['ID']))

#ABAS
df_planilha = pd.read_excel("TESTE.xlsx", sheet_name=None)
df_produtos = df_planilha['PRODUTOS']
df_compras_produtos = df_planilha['COMPRAS PRODUTOS']
df_vendas_produtos = df_planilha['VENDAS PRODUTOS']

#print(f"\nColunas disponíveis em PRODUTOS: {list(df_produtos.columns)}")
#print(f"Primeiras 5 linhas da aba PRODUTOS:")
#print(df_produtos.head())

#Preenche ID-CFOP baseado no CFOP
nao_encontrados = None

if 'CFOP' in df_produtos.columns and 'ID-CFOP' in df_produtos.columns:
    # Mapeia CFOP para ID usando o dicionário
    df_produtos['ID-CFOP'] = df_produtos['CFOP'].map(cfop_to_id).fillna('')
    
    #print(f"\nResultado após preenchimento:")
    #print(df_produtos[['CFOP', 'ID-CFOP']].head(10))
    
    # Verifica se há CFOPs não encontrados
    nao_encontrados = df_produtos[df_produtos['ID-CFOP'] == '']
    if not nao_encontrados.empty:
        print("\nCFOPs não encontrados na matriz tributária:")
        print(nao_encontrados['CFOP'].dropna().unique())
else:
    print("Colunas 'CFOP' ou 'ID-CFOP' não encontradas na aba PRODUTOS")

if 'CFOP' in df_vendas_produtos.columns and 'TIPO CFOP' in df_vendas_produtos.columns:
    teste = df_vendas_produtos['TIPO CFOP'] = df_vendas_produtos['CFOP'].map(cfop_to_id)

if 'CFOP' in df_compras_produtos.columns and 'TIPO CFOP' in df_compras_produtos.columns:
    teste = df_compras_produtos['TIPO CFOP'] = df_compras_produtos['CFOP'].map(cfop_to_id)

#Vetorizando os produtos
vectorizer = TfidfVectorizer()
X_train_vec = vectorizer.fit_transform(X_train)

#Criando o modelo de treino
model = RandomForestClassifier()
model.fit(X_train_vec, y_train)

#Aplicando o modelo
produtos_teste = df_produtos['PRODUTOS'].astype(str)
X_test_vec = vectorizer.transform(produtos_teste)

probas = model.predict_proba(X_test_vec)
max_probas = probas.max(axis=1)

#Garantindo uma qualidade sobre a decisão
limiar = 0.6
df_produtos['CLASSIFICAÇÃO'] = model.predict(X_test_vec)
df_produtos['CONFIANCA'] = max_probas
df_produtos.loc[df_produtos['CONFIANCA'] < limiar, 'CLASSIFICAÇÃO'] = 'Não Classificado'

#Checando os produtos não classificados
mask_nc = df_produtos['CLASSIFICAÇÃO'] == 'Não Classificado'
produtos_nc = df_produtos.loc[mask_nc, 'PRODUTOS'].astype(str).tolist()

X_test_vec = vectorizer.transform(df_compras_produtos['Nome Produto'].astype(str))
probas = model.predict_proba(X_test_vec)
max_probas = probas.max(axis=1)

#Garantindo uma qualidade sobre a decisão
limiar = 0.6
df_compras_produtos['TIPO CLASSIFICAÇÃO'] = model.predict(X_test_vec)
df_compras_produtos['CONFIANCA'] = max_probas
df_compras_produtos.loc[df_compras_produtos['CONFIANCA'] < limiar, 'TIPO CLASSIFICAÇÃO'] = 'Não Classificado'
df_compras_produtos.loc[df_compras_produtos['CFOP'].astype(str) == 1556, 'CLASSIFICAÇÃO'] = 'USO E CONSUMO'

# Tratar CFOP — converte para numérico com segurança e aplica override
mask_cfop = pd.to_numeric(df_compras_produtos['CFOP'], errors='coerce') == 1556
df_compras_produtos.loc[mask_cfop, 'TIPO CLASSIFICAÇÃO'] = 'USO E CONSUMO'
df_compras_produtos.loc[mask_cfop, 'CONFIANCA'] = 1.0

X_test_vec = vectorizer.transform(df_vendas_produtos['Nome Produto'].astype(str))
probas = model.predict_proba(X_test_vec)
max_probas = probas.max(axis=1)

#Garantindo uma qualidade sobre a decisão
limiar = 0.6
df_vendas_produtos['TIPO CLASSIFICAÇÃO'] = model.predict(X_test_vec)
df_vendas_produtos['CONFIANCA'] = max_probas
df_vendas_produtos.loc[df_vendas_produtos['CONFIANCA'] < limiar, 'TIPO CLASSIFICAÇÃO'] = 'Não Classificado'

#Tratar CFOP — converte para numérico com segurança e aplica override
mask_cfop = pd.to_numeric(df_vendas_produtos['CFOP'], errors='coerce') == 1556
df_vendas_produtos.loc[mask_cfop, 'TIPO CLASSIFICAÇÃO'] = 'USO E CONSUMO'
df_vendas_produtos.loc[mask_cfop, 'CONFIANCA'] = 1.0

#Prepara lista de CFOPs não classificados com segurança
cfops_nc = []
if isinstance(nao_encontrados, type(None)) or nao_encontrados.empty:
    cfops_nc = []
else:
    #Converte para string para evitar problemas ao escrever
    cfops_nc = [str(x) for x in nao_encontrados['CFOP'].dropna().unique()]

with open('produtos_nao_classificados.txt', 'w', encoding='utf-8') as f:
    f.write("PRODUTOS NÃO CLASSIFICADOS\n")
    for p in produtos_nc:
        f.write(p + '\n')

    f.write("\nCFOP NÃO CLASSIFICADOS\n")
    for i in cfops_nc:
        f.write(i + '\n')
    
df_produtos.loc[df_produtos['CFOP'] == 1556, 'CLASSIFICAÇÃO'] = 'USO E CONSUMO'

# Carrega o workbook existente
wb = load_workbook("TESTE.xlsx", data_only=False)  # data_only=False mantém fórmulas
ws = wb["PRODUTOS"]
ws_compras_produtos = wb["COMPRAS PRODUTOS"]
ws_vendas_produtos = wb["VENDAS PRODUTOS"]

# Mapeia colunas por nome (assumindo cabeçalhos na primeira linha)
headers = [cell.value for cell in ws[1]]
col_idx = {name: i+1 for i, name in enumerate(headers)}

# Garante que colunas existam (cria cabeçalhos se faltarem)
def ensure_col(name):
    if name not in col_idx:
        ws.cell(row=1, column=ws.max_column+1, value=name)
        col_idx[name] = ws.max_column
ensure_col("CLASSIFICAÇÃO")
ensure_col("CONFIANCA")
ensure_col("ID-CFOP")

# Escreve valores linha a linha (assumindo que df_produtos está alinhado à aba)
for i, row in df_produtos.iterrows():
    excel_row = i + 2  # +2 por causa do cabeçalho 1-based
    ws.cell(row=excel_row, column=col_idx["CLASSIFICAÇÃO"], value=row["CLASSIFICAÇÃO"])
    ws.cell(row=excel_row, column=col_idx["CONFIANCA"], value=float(row["CONFIANCA"]))
    ws.cell(row=excel_row, column=col_idx["ID-CFOP"], value=str(row["ID-CFOP"]))

#COMPRAS PRODUTOS
headers = [cell.value for cell in ws_compras_produtos[1]]
col_idx = {name: i+1 for i, name in enumerate(headers)}

for i, row in df_compras_produtos.iterrows():
    excel_row = i + 2
    ws_compras_produtos.cell(row=excel_row, column=col_idx["TIPO CFOP"], value=row["TIPO CFOP"])
    ws_compras_produtos.cell(row=excel_row, column=col_idx["TIPO CLASSIFICAÇÃO"], value=row["TIPO CLASSIFICAÇÃO"])

#VENDAS PRODUTOS
headers = [cell.value for cell in ws_vendas_produtos[1]]
col_idx = {name: i+1 for i, name in enumerate(headers)}

for i, row in df_vendas_produtos.iterrows():
    excel_row = i + 2
    ws_vendas_produtos.cell(row=excel_row, column=col_idx["TIPO CFOP"], value=row["TIPO CFOP"])
    ws_vendas_produtos.cell(row=excel_row, column=col_idx["TIPO CLASSIFICAÇÃO"], value=row["TIPO CLASSIFICAÇÃO"])

# Salva sem reescrever as outras abas (fórmulas preservadas)
wb.save("TESTE.xlsx")